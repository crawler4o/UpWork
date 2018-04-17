import csv
from openpyxl import load_workbook


def xlsx_writer(sheet, csv_file):
    with open(csv_file, 'r') as open_file:
        csv_data = csv.reader(open_file)
        row_track = 1
        column_track = 1
        for row_ in csv_data:
            for column_ in row_:
                sheet.cell(row=row_track, column=column_track).value = column_
                column_track +=1
            row_track +=1
            column_track = 1


if __name__ == '__main__':
    wb = load_workbook('Master_Data.xlsx')
    movies = wb['Movies']
    music = wb['Music']
    plays = wb['Play']

    xlsx_writer(movies, 'movies.csv')
    xlsx_writer(music, 'music.csv')
    xlsx_writer(plays, 'plays.csv')

    wb.save('Master_Data.xlsx')
