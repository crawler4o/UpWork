import shutil, csv, traceback
from openpyxl import load_workbook



rnd_file_name = 'Released plan B1, B2, B3 2018.05.08.xlsx' 
output_files_list = 'list.txt'
ss_template_file = 'New Site Solution Template V3.91 Template updated.xlsx'
ss_template_file_dummy = 'Dummy_template.xlsx'
inv_file = 'Inventory_Board_20180511_180809.csv'
inv_cabinets = 'Inventory_Cabinet_20180511_180938.csv'
inv_antennas = 'Inventory_Antenna_20180511_180750.csv'
inv_subracks = 'Inventory_Subrack_20180511_180930.csv'
rnp_summary_file = 'MV_RNP_SUMMARY_20180508.csv'
nis_exp = 'NISRV_07.05.2018.csv'



def rnd_shot(ss, job_on): # This will paste the RND 
	wb = load_workbook(filename = rnd_file_name) # opening the rnd
	sswb = load_workbook(filename = ss) # opening the ss file
	sheet: object = wb['Running cells'] #defining the rnd sheet to read
	sheet_ss = sswb['Front Page'] # identifying the ss sheet to write on
	d_row = 3 # ss row to write the rnd data on
	for x in range(2, 3000):
		job = 'M'+ str(x) # row of the job in the rnd
		r_job = sheet[job].value # the job id from the rnd
		if job_on == r_job: # if the job name in the solution file name equals one on the rnd
			# print("MATCH!") # Troubleshooting entry - Prints match when a line corresponding to the current ss file is found.
			for col in range(1, 20):
				cp_s = sheet.cell(column = col, row = x ).value # source cell for copying
				if not cp_s:
					cp_s = ''
				sheet_ss.cell(column = col+7, row = d_row, value = "{0}".format(cp_s))		# writing the value to the destination cell
			d_row +=1 # ss going to write on the next row
	
	sswb.save(filename = ss)


	
def parse_inv(inv_file, site_id):
	bbu_boards = ['UPEU', 'UEIU', 'WBBP', 'LBBP', 'UBBP', 'UMPT', 'WMPT', 'LMPT', 'GTMU', 'UTRP'] # list of wanted boards to be kept in the BBU boards dictionary
	radio_boards = ['MRFU', 'WRFU', 'LRFU', 'MRRU', 'LRRU'] # list or boards to be kept in the RF boards dictionary
	
	with open(inv_file, 'r', encoding ='utf-8') as csv_file:
		brd_inv = csv.reader(csv_file)
		node = {}
		for row in brd_inv: # add the boards that correspond to the site
			if site_id in row[2]:
				node[row[23]] = [row[10], row[25], row[5], row[6], row[2][7]] # {SerialNo:[sub-rack,slot, board name, bord type(bom), U/L ( reporting technology )]}
			
	to_del = [] # list of entries to be removed
	node_rf = {} # dictionary for the rf boards
	node_bbu = {}
	
	for x in node: # identify the boards
		if node[x][2] in radio_boards: # add the Rf boards to the node_rf dictionary
			node_rf[x] = node[x]
		
		if node[x][2] in bbu_boards:
			node_bbu[x] = node[x]
	
	node_bbu = convert_dic(node_bbu) # convert the dictionary for BBU . one more item is added after the procedure {SerialNo:[sub-rack,slot, board name, bord type(bom), U/L ( reporting technology ), row on SS ]}
	node_rf = convert_dic(node_rf) # convert the dictionary for RF one more item is added after the procedure {SerialNo:[sub-rack,slot, board name, bord type(bom), U/L ( reporting technology ), row on SS ]}
	
	
	# updating the BBU dictionaries in case there are two BBUs added as Subrack 0 on site.
	
	there_already7 = False
	there_already6 = False
	double_trouble = False
	for x in node_bbu: # checks for duplicated boards on slots 6 & 7, assuming that having a board on one of there is a must.
		if node_bbu[x][1] == '7' and node_bbu[x][0] == '0':
			if not there_already7:
				there_already7 = True
			else:
				double_trouble = True
			
		if node_bbu[x][1] == '6' and node_bbu[x][0] == '0':
			if not there_already6:
				there_already6 = True
			else:
				double_trouble = True
	
	if double_trouble: # updates the sub-rack number of the LTE boards to '1' in case of duplicated 
		for x in node_bbu:
			if node_bbu[x][4] == 'L':
				node_bbu[x][0] = '1'
	
	
	
	return [node_bbu, node_rf]

	

def write_boards(nodes, file): # writes the third field of the nodes dictionary ( now it's BOM, value 4 of the dict should be the RRU row ) 
	sswb = load_workbook(filename = file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on
	for x in nodes[0]: # checking and writing the BBU boards
		if nodes[0][x][0] == '0': # BBU in subrack 0
			if int(nodes[0][x][1]) < 8: # to write the boards on the first 7 slots
				if not sheet_ss.cell(column = 3, row = 22 +int(nodes[0][x][1])).value: # make sure that the field is empty
					sheet_ss.cell(column = 3, row = 22 +int(nodes[0][x][1]), value = "{0}".format(nodes[0][x][3]))
					sheet_ss.cell(column = 4, row = 22 +int(nodes[0][x][1]), value = 1)
				else:
					sheet_ss.cell(column = 3, row = 22 +int(nodes[0][x][1]), value = 'error')		
			else: #to write the UPEU and UEIU
				if not sheet_ss.cell(column = 3, row = 12 +int(nodes[0][x][1])).value: # make sure that the field is empty
					sheet_ss.cell(column = 3, row = 12 +int(nodes[0][x][1]), value = "{0}".format(nodes[0][x][3]))
					sheet_ss.cell(column = 4, row = 12 +int(nodes[0][x][1]), value = 1)
				else:
					sheet_ss.cell(column = 3, row = 12 +int(nodes[0][x][1]), value = 'error')
		else: # BBU in subrack 1
			if int(nodes[0][x][1]) < 8: # to write the boards on the first 7 slots
				if not sheet_ss.cell(column = 6, row = 22 +int(nodes[0][x][1])).value: # make sure that the field is empty
					sheet_ss.cell(column = 6, row = 22 +int(nodes[0][x][1]), value = "{0}".format(nodes[0][x][3]))
					sheet_ss.cell(column = 7, row = 22 +int(nodes[0][x][1]), value = 1)
				else:
					sheet_ss.cell(column = 6, row = 22 +int(nodes[0][x][1]), value = 'error')	
			else: #to write the UPEU and UEIU
				if not sheet_ss.cell(column = 6, row = 12 +int(nodes[0][x][1])).value: # make sure that the field is empty
					sheet_ss.cell(column = 6, row = 12 +int(nodes[0][x][1]), value = "{0}".format(nodes[0][x][3]))
					sheet_ss.cell(column = 7, row = 12 +int(nodes[0][x][1]), value = 1)
				else:
					sheet_ss.cell(column = 6, row = 12 +int(nodes[0][x][1]), value = 'error')

	for x in nodes[1]: # checking and writing the RF boards

		if not sheet_ss.cell(column = 3, row = int(nodes[1][x][5])).value: # if the field is empty
			sheet_ss.cell(column = 3, row = int(nodes[1][x][5]), value = "{0}".format(nodes[1][x][3]))
			sheet_ss.cell(column = 4, row = int(nodes[1][x][5]), value = 1)
		elif sheet_ss.cell(column = 3, row = int(nodes[1][x][5])).value == nodes[1][x][3]: #if the same is already added
			sheet_ss.cell(column = 4, row = int(nodes[1][x][5]), value = int(sheet_ss.cell(column = 4, row = int(nodes[1][x][5])).value) + 1)
		elif sheet_ss.cell(column = 3, row = int(nodes[1][x][5])).value != nodes[1][x][3] and not sheet_ss.cell(column = 6, row = int(nodes[1][x][5])).value: # if different in column 1 and nothing in column 2
			sheet_ss.cell(column = 6, row = int(nodes[1][x][5]), value = "{0}".format(nodes[1][x][3]))
			sheet_ss.cell(column = 7, row = int(nodes[1][x][5]), value = 1)
		elif sheet_ss.cell(column = 3, row = int(nodes[1][x][5])).value != nodes[1][x][3] and sheet_ss.cell(column = 6, row = int(nodes[1][x][5])).value == nodes[1][x][3]: # if different in column 1 and same in column 2
			sheet_ss.cell(column = 7, row = int(nodes[1][x][5]), value = int(sheet_ss.cell(column = 7, row = int(nodes[1][x][5])).value) + 1)
		else: # if different in both column 1 and 2 - return an error
			sheet_ss.cell(column = 3, row = int(nodes[1][x][5]), value = 'error')
				
	sswb.save(filename = file)

	
	
def convert_dic(dic): # convert the BOM codes into names as per the below dictionary. Adds row for the RF modules
	
	translate = { 
			'QWL3WBBPF3':['WBBPf3', '0'],
			'WD22UMPTb1':['UMPTb1', '0'],
			'WD2MUPEUC':['UPEUc', '0'],
			'WD5MJFUGG8E':['MRFUd (900)', '35'],
			'QWL1WBBPD2':['WBBPd2', '0'],
			'WD22UMPTa2':['UMPTa2', '0'],
			'WD2M1UEIU':['UEIU', '0'],
			'WD5MJFUBG8E':['MRFUd (900)', '35'],
			'QWL3WBBPF3':['WBBPf3', '0'],
			'WD22WMPT':['WMPT', '0'],
			'WD5MIFUBC10':['WRFUd (2100 2T2R)', '37'],
			'WD2MUPEUD2':['UPEUd', '0'],
			'WD5MJRUA880':['RRU3928 (900) old', '35'],
			'QWL1WBBPD1':['WBBPd1', '0'],
			'WD5MMRFU78':['MRFU (900) old', '35'],
			'WD22UBBPd1':['UBBPd1', '0'],
			'WD5MZAAZGAF':['3965d (800&900)', '34'],
			'WD22UBBPd4':['UBBPd4', '0'],
			'WD22UBBPd3':['UBBPd3', '0'],
			'WD5MZAAZGAFX':['RRU3965 (800&900)', '34'],
			'WD5MJRUE88E':['MRRU3938  (900)', '35'],
			'QWL1WBBPD3':['WBBPd3', '0'],
			'WD5MIRUD810':['RRU3838 2T2R(2100)', '37'],
			'QWM2UTRP4':['UTRP4', '0'],
			'WD22UMPTb2':['UMPTb2', '0'],
			'WD5MIRUA810':['RRU3828 2T2R(2100) old', '37'],
			'WD5MWFUB81':['MARP 2100', '37'],
			'WD5MJRUYCY0':['RRU 3961(800&900) old', '34'],
			'WD22UBBPd6':['UBBPd6', '0'],
			'WD2MWRFU81':['WRFUd(2100 2T2R) old', '37'],
			'QWL3WBBPF1':['WBBPf1', '0'],
			'QWL1WBBPF4':['WBBPf4', '0'],
			'WD5MMRFU78B':['MRFUv2 (900)', '35'],
			'WD3M1RRU4':['RRU3801E (2100)', '37'],
			'WD5MARU261':['RRU3804 (2100)', '37'],
			'WD5MIRUDC10':['RRU3839 2T2R(2100) new', '37'],
			'WD2MUPEUA':['UPEUa', '0'],
			'WD5MLFUHCK0':['LRFUe(800)', '34'],
			'WD5MJFUBG30':['MRFUd(1800)', '36'],
			'WD23LBBPD1':['LBBPd1', '0'],
			'WD5MIFUBCK0':['LRFUe(800)', '34'],
			'WD5MJFUGG30':['MRFUd(1800)', '36'],
			'WD22LMPT1':['LMPT', '0'],
			'WD22LBBPC':['LBBPc', '0'],
			'WD5MJFUHG30':['MRFUd(1800)', '36'],
			'WD5MIRUB8KA':['RRU3220 (800) old', '34'],
			'WD5MJRUA830':['RRU3928 (1800) old', '36'],
			'WD5MLRUH870':['RRU3268 2T2R(2600)', '38'],
			'WD5MJRUE830':['RRU3938(1800 2T2R)', '36'],
			'WD5MLRUA8K0':['LRRU3268(800)', '34'],
			'WD5MJRUIG30':['RRU3971 4T4R(1800)', '36'],
			'WD5MLRUYG70':['RRU3281 4T4R(2600)', '38'],
			'WD5MLFUHC70':['LRFU 2T2R(2600)', '38'],
			'WD5MLFU287C':['MARP(2600)', '38'],
			'WD5MLRUC870':['RRU3240(2600 2T4R) old', '38'],
			'WD5MLRUE870':['RRU3260(2600 2T4R) old', '38'],
			'WD22LBBPD1':['LBBPd1', '0'],
			'WD5MLRUA8K0L':['RRU3268 2T2R(2600)', '38'],
			'WD22LBBPD3':['LBBPd3', '0'],
			'WD22LBBPD2':['LBBPd2', '0'],
			'WD5MIRU187C':['RRU3201(2600) old', '38'],
			'WD5MMRFU73B':['MRFUv2 (1800)', '36'],
			'WD22GTMUb':['GTMUb', '0'],
		}
	
	for x in dic:
		dic[x].append(translate[dic[x][3]][1]) # adds the dummy_SS row on which the info should be written
		dic[x][3] = translate[dic[x][3]][0] # replaces the BOM with name based on 'board type'
		
	return dic

	
	
def add_ant(file, site_id, dummy_file): # adds the antennas to the dummy solution
	
	antennas = {} # dictionary to store the antennas.

	# Read the antennas from the RNP export and add the info to {}
	with open(file, 'r', encoding ='utf-8') as csv_file:	# read from the RNP export
		rnp_data = csv.reader(csv_file, delimiter=';')
		for row in rnp_data:
			if site_id in row[0]:
				if row[2] not in antennas.keys() or antennas[row[2]] == row[16]: # add to dictionary with clean A B or C key for first layer antennas
					antennas[row[2]] = row[16]
				elif antennas[row[2]] != row[16] and row[2]+'2' not in antennas.keys(): # add to dictionary with A2, B2, C2 for second layer antennas
					antennas[row[2]+'2'] = row[16]
				else: # adds 'error' value if more than 2 layers of antennas exist
					antennas[row[2]] = 'error'
	
		
	# Write the {} to the dummy solution
			
	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on	
		
	for x in antennas:
		if x == 'A':
			sheet_ss.cell(column = 3, row = 45, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 45, value = 1)
		elif x == 'B':
			sheet_ss.cell(column = 3, row = 47, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 47, value = 1)
		elif x ==  'C':
			sheet_ss.cell(column = 3, row = 49, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 49, value = 1)
		elif x == 'D':
			sheet_ss.cell(column = 3, row = 51, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 51, value = 1)
		elif x == 'A2':
			sheet_ss.cell(column = 3, row = 46, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 46, value = 1)
		elif x == 'B2':
			sheet_ss.cell(column = 3, row = 48, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 48, value = 1)
		elif x == 'C2':
			sheet_ss.cell(column = 3, row = 50, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 50, value = 1)
		elif x == 'D2':
			sheet_ss.cell(column = 3, row = 52, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 52, value = 1)

	sswb.save(filename = dummy_file)

	

def add_cab(file, site_id, dummy_file):
	
	def dist_cab(column_, cabinet, sheet_ss): # adds the given cabinet to the sheet
			
		for x in range(4):
			if sheet_ss.cell(column = column_, row = 5 + x).value: # if a cabinet is already added
				if x == 4:
					sheet_ss.cell(column = column_  , row = 5 + x, value = "{0}".format('error'))
					break
				else:
					pass
			else: # if it is empty
				sheet_ss.cell(column = column_, row = 5 + x, value = "{0}".format(cabinet))
				sheet_ss.cell(column = column_ + 1 , row = 5 + x, value = "{0}".format(1))
				break
			
		return sheet_ss
	
		
	cabinets = {} # The dictionary
	with open(file, 'r', encoding ='utf-8') as csv_file: # read the cabinet export
		cab_data = csv.reader(csv_file)
		no_ser = 1
		for row in cab_data:	
			if site_id in row[2]: # if the site ID is in the NE name
				if len(row[15]) > 2: # if 'SN(BarCode)' is not empty
					cabinets[row[15]] = row[14] # adds the value of 'Rack Type' with 'SN(BarCode)' as a key
				else: # if 'SN(BarCode)' is empty
					cabinets[no_ser] = row[14]+'_NS'
					no_ser += 1
	
	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on	
	
	
	for x in cabinets:
		if  cabinets[x] == 'RFC' or cabinets[x] == 'BBC' or cabinets[x] == 'RFC_NS' or cabinets[x] == 'BBC_NS':
			sheet_ss = dist_cab(6, cabinets[x], sheet_ss)
		else:
			sheet_ss = dist_cab(3, cabinets[x], sheet_ss)

	
	sswb.save(filename = dummy_file) # save and close the dummy file
	

		
def add_ret(file, site_id, dummy_file): # will use the 'Vendor Name' and 'Vendor Unit Family Type' fields to count + serial number as a key.
	rets = [] # The list
	with open(file, 'r', encoding ='utf-8') as csv_file: # read the antenna export
		ret_data = csv.reader(csv_file)
		for row in ret_data:
			if site_id in row[2] and row[19] == 'SINGLE_RET': # if the site ID is in the NE name
 				rets.append(row[18])

	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on	
	
	if rets.count('KA') > 0: 
		sheet_ss.cell(column = 4, row = 76, value = rets.count('KA'))
		sheet_ss.cell(column = 3, row = 76, value = "{0}".format('KA RET'))	
		
	sswb.save(filename = dummy_file)			
	


def add_tma(file, site_id, dummy_file):
	tmas = {} # key is the serial number
	with open(file, 'r', encoding = 'utf-8') as csv_file: # read the antenna export
		tma_data = csv.reader(csv_file)
		for row in tma_data:
			if site_id in row[2] and row[19] == 'TMA': # if the site ID is in the NE name
				if len(row[11]) > 2: # if the TMA has a BOM code 
					tmas[row[15]] = [row[11], row[2][7]] # {Serial:[bom, U/L]}
				else:
					tmas[row[15]] = [row[14], row[2][7]] #  {serial:[antenna model, U/L}
	
	tma_translation_dic = { # Dictionary to convert the BOMs to names
					'27100072':'ATADU2001 (800&900)',
					'27100046':'STMA 2100',
					'27100074':'ATADU2002 (1800&2100)',
					'27100073':'ATADU2005 (800&900) 2IN-4OUT',
					'27100075':'ATADU2003 (1800+2100) 2IN-4OUT',
					'27100083':'ATADU2002 (1800&2100)',
					'27100045':'STMA 2100',
					'99044AHL':'STMA 2100',
					'27100060':'STMA 1800',
					'27100052':'ATA262000DTMA (2600)',
					'27100112':'ATADU2015 MTMA 1800&2100+2600 2IN-4OUT',
					'27100037':'STMA 1800',
					'DTMA2100':'STMA 2100',
					'78210517':'KA 78210517 ( 800 & 900 ,2in4out)',
					'DTMA800':'DTMA800',
					'DTMA2600':'ATA262000DTMA (2600)',
					'DTMA800&900':'DTMA800&900 (no BOM)',
					'DTMA1800':'STMA 1800',
					'DTMA1800&2100':'DTMA1800&2100 (no BOM)',
					}
		
	
	# convert what is on the TMA dict ( bom or ant model ) to the target values
	for x in tmas:
		tmas[x] = [tma_translation_dic[tmas[x][0]], tmas[x][1]] 
	
	
	
	# remove the 2in4out tmas reported by the 3G, as every TMA appears as 2 in the export.
	# First identify the 2in4outs reported by the 3G
	for x in tmas:
		if (tmas[x][0] == 'ATADU2003 (1800+2100) 2IN-4OUT' or tmas[x][0] == 'ATADU2005 (800&900) 2IN-4OUT') and tmas[x][1] == 'U':
			tmas[x][0] = 'del'


	# Now update the dictionary
	tmas = {k:v for k,v in tmas.items() if v[0] !='del'}
	
			
		
	# filling the dummy_SS
	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on	
	
	tmas_reached_ant_no = [] # adding to this list TMAs that have reached the number of sectors
	ant_number = 0
	
	for x in range(7): # counting the antennas
		if sheet_ss.cell(column = 3, row = 45+x).value:
			ant_number +=1
	
	for x in tmas: # filling the SS
		pos1 = 53
		pos2 = 56
		pos3 = 59
		pos4 = 62
		for z in range(3): # checking if this model TMA is the first, second, third or fourth on the site
			if sheet_ss.cell(column = 3, row = pos1).value != tmas[x][0] and sheet_ss.cell(column = 3, row = pos1).value: 
				pos1 +=1
				pos2 +=1
				pos3 +=1
				pos4 +=1
		

		for z in range(3): # checking if the tma to be added belongs to a second layer and the next field is free
			if tmas[x][0] in tmas_reached_ant_no and not sheet_ss.cell(column = 3, row = pos1+1+z).value: 
				pos1 +=1+z
				pos2 +=1+z
				pos3 +=1+z
				pos4 +=1+z
				break
		
		
		
		
		if sheet_ss.cell(column = 3, row = pos1).value == tmas[x][0]: # check if the same is already filled in somewhere
			if sheet_ss.cell(column = 3, row = pos2).value == tmas[x][0]:
				if sheet_ss.cell(column = 3, row = pos3).value == tmas[x][0]:
					if sheet_ss.cell(column = 3, row = pos4).value == tmas[x][0]:
						sheet_ss.cell(column = 3, row = pos4, value = "{0}".format('error'))
					else:
						sheet_ss.cell(column = 3, row = pos4, value = "{0}".format(tmas[x][0]))
						sheet_ss.cell(column = 4, row = pos4, value = 1)
						tmas_reached_ant_no.append(tmas[x][0]) # we are not working with more than 4 sectors, so no way to continue without shifting down
				else:
					sheet_ss.cell(column = 3, row = pos3, value = "{0}".format(tmas[x][0]))
					sheet_ss.cell(column = 4, row = pos3, value = 1)
					if ant_number == 3:
						tmas_reached_ant_no.append(tmas[x][0]) # to shift, assuming that we have 2 TMAs of kind here
			else:
				sheet_ss.cell(column = 3, row = pos2, value = "{0}".format(tmas[x][0]))
				sheet_ss.cell(column = 4, row = pos2, value = 1)
				if ant_number == 2:
					tmas_reached_ant_no.append(tmas[x][0]) # to shift, assuming that we have 2 TMAs of kind here
		elif not sheet_ss.cell(column = 3, row = pos1).value: # if it's empty
			sheet_ss.cell(column = 3, row = pos1, value = "{0}".format(tmas[x][0]))
			sheet_ss.cell(column = 4, row = pos1, value = 1)
			if ant_number == 1:
				tmas_reached_ant_no.append(tmas[x][0]) # to shift, assuming that we have 2 TMAs of kind here
	
	
	sswb.save(filename = dummy_file)						

	

def add_feeder(file, site_id, dummy_file):
	
	feeders = {} # dictionary to store the antennas.

	# Read the feeders from the RNP export and add the info to {}
	with open(file, 'r', encoding ='utf-8') as csv_file:	# read from the RNP export
		rnp_data = csv.reader(csv_file, delimiter=';')
		for row in rnp_data:
			if site_id in row[0] and row[20]: # if the site id matches and feeder length is given
				if row[2]+'1' not in feeders.keys(): # add to dictionary with key A1, B1 or C1 key for first layer feeders
					feeders[row[2]+'1'] = [row[20], row[19], row[4]]
				elif row[2]+'2' not in feeders.keys(): # add to dictionary with key A2, B2 or C2 key for second layer feeders
					if row[4] != feeders[row[2]+'1'][2]: # if there is no feeder added already for this technology
						feeders[row[2]+'2'] = [row[20], row[19], row[4]]
					else:
						continue
				elif row[2]+'3' not in feeders.keys(): # add to dictionary with key A3, B3 or C3 key for third layer feeders
					if row[4] != feeders[row[2]+'1'][2] and row[4] != feeders[row[2]+'2'][2]: 
						feeders[row[2]+'3'] = [row[20], row[19], row[4]]
					else:
						continue
				elif row[2]+'4' not in feeders.keys(): # add to dictionary with key A4, B4 or C4 key for third layer feeders
					if row[4] != feeders[row[2]+'1'][2] and row[4] != feeders[row[2]+'2'][2] and row[4] != feeders[row[2]+'3'][2]: 
						feeders[row[2]+'4'] = [row[20], row[19], row[4]]
					else:
						continue
				elif row[2]+'5' not in feeders.keys(): # add to dictionary with key A5, B5 or C5 key for third layer feeders
					if row[4] != feeders[row[2]+'1'][2] and row[4] != feeders[row[2]+'2'][2] and row[4] != feeders[row[2]+'3'][2] and row[4] != feeders[row[2]+'4'][2]:
						feeders[row[2]+'5'] = [row[20], row[19], row[4]]
					else:
						continue
				elif row[2]+'5' not in feeders.keys(): # add to dictionary with key A6, B6 or C6 key for third layer feeders
					if row[4] != feeders[row[2]+'1'][2] and row[4] != feeders[row[2]+'2'][2] and row[4] != feeders[row[2]+'3'][2] and row[4] != feeders[row[2]+'4'][2] and row[4] != feeders[row[2]+'5'][2]:
						feeders[row[2]+'6'] = [row[20], row[19], row[4]]
					else:
						continue		
				else: # adds 'error' value if more than 6 layers of feeders exist
					if row[4] != feeders[row[2]+'1'][2] and row[4] != feeders[row[2]+'2'][2] and row[4] != feeders[row[2]+'3'][2] and row[4] != feeders[row[2]+'4'][2] and row[4] != feeders[row[2]+'5'][2] and row[4] != feeders[row[2]+'6'][2]: 						
						feeders[row[2]+'6'] = 'error'
					else:
						continue
	
	# Write the {} to the dummy solution
			
	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on	
		
	for x in feeders:
		if x == 'A1':
			sheet_ss.cell(column = 3, row = 78, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 78, value = 1)
		elif x == 'B1':
			sheet_ss.cell(column = 3, row = 81, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 81, value = 1)
		elif x == 'C1':
			sheet_ss.cell(column = 3, row = 84, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 84, value = 1)
		elif x == 'D1':
			sheet_ss.cell(column = 3, row = 87, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 87, value = 1)
		elif x == 'A2':
			sheet_ss.cell(column = 3, row = 79, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 79, value = 1)
		elif x == 'B2':
			sheet_ss.cell(column = 3, row = 82, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 82, value = 1)
		elif x == 'C2':
			sheet_ss.cell(column = 3, row = 85, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 85, value = 1)
		elif x == 'D2':
			sheet_ss.cell(column = 3, row = 88, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 88, value = 1)
		elif x == 'A3':
			sheet_ss.cell(column = 3, row = 80, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 80, value = 1)
		elif x == 'B3':
			sheet_ss.cell(column = 3, row = 83, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 83, value = 1)
		elif x == 'C3':
			sheet_ss.cell(column = 3, row = 86, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 86, value = 1)
		elif x == 'D3':
			sheet_ss.cell(column = 3, row = 89, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 4, row = 89, value = 1)
		elif x == 'A4':
			sheet_ss.cell(column = 6, row = 78, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 78, value = 1)
		elif x == 'B4':
			sheet_ss.cell(column = 6, row = 81, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 81, value = 1)
		elif x == 'C4':
			sheet_ss.cell(column = 6, row = 84, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 84, value = 1)
		elif x == 'D4':
			sheet_ss.cell(column = 6, row = 87, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 87, value = 1)
		elif x == 'A5':
			sheet_ss.cell(column = 6, row = 79, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 79, value = 1)
		elif x == 'B5':
			sheet_ss.cell(column = 6, row = 82, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 82, value = 1)
		elif x == 'C5':
			sheet_ss.cell(column = 6, row = 85, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 85, value = 1)
		elif x == 'D5':
			sheet_ss.cell(column = 6, row = 88, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 88, value = 1)
		elif x == 'A6':
			sheet_ss.cell(column = 6, row = 80, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 80, value = 1)
		elif x == 'B6':
			sheet_ss.cell(column = 6, row = 83, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 83, value = 1)
		elif x == 'C6':
			sheet_ss.cell(column = 6, row = 86, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 86, value = 1)
		elif x == 'D6':
			sheet_ss.cell(column = 6, row = 89, value = '{0}m, {1}, {2}'.format(feeders[x][0], feeders[x][1], feeders[x][2]))
			sheet_ss.cell(column = 7, row = 89, value = 1)					
			
	sswb.save(filename = dummy_file)
	
	
	
def add_owner_and_nis(file, site_id, dummy_file):
	
	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Front Page'] # identifying the ss sheet to write on	
	
	with open(file, 'r', encoding ='ansi') as csv_file: # read from the NISRV export
		owner_data = csv.reader(csv_file, delimiter=';')
		for row in owner_data:
			if site_id in row[1]:
				sheet_ss.cell(column = 3, row = 32, value = '{}'.format(row[2])) # owner
				sheet_ss.cell(column = 3, row = 28, value = '{}'.format(row[12])) # nis version
				sheet_ss.cell(column = 3, row = 29, value = '{}'.format(row[13])) # nis date
	
	sswb.save(filename = dummy_file)

	
	
def add_bbu(file, site_id, dummy_file):

	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on	
	
	bbus = {}
	with open(file, 'r', encoding ='utf-8') as csv_file: # read from the NISRV export
		subrack_data = csv.reader(csv_file, delimiter=',')
		for row in subrack_data:
			if site_id in row[2] and (row[7] == '0' or row[7] == '1'):
				bbus[row[18]] = [row[8], row[7]] # { serial:[frame type, subrack no]}
				
		offset = 0		
	for x in bbus:
		offset += 3
		if offset <7:
			sheet_ss.cell(column = offset, row = 21, value = '{}'.format(bbus[x][0]))
			sheet_ss.cell(column = offset+1, row = 21, value = '1')
		else:
			sheet_ss.cell(column = 6, row = 21, value = 'too many BBUs')
			sheet_ss.cell(column = 7, row = 21, value = '2')			

	
	sswb.save(filename = dummy_file)
	


def add_combiners(dummy_file): # adding combiners based on RFUs and number of TMAs. We assume that if there is an RFU, there should be a TMA.
	
	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on	

	tma_lb = ['ATADU2001 (800&900)', 'ATADU2005 (800&900) 2IN-4OUT', 'DTMA800&900 (no BOM)', 'KA 78210517 ( 800 & 900 ,2in4out)', 'DTMA800 (no BOM)']
	tma_hb = ['STMA 2100', 'ATADU2002 (1800&2100)', 'ATADU2003 (1800+2100) 2IN-4OUT', 'STMA 1800', 
			'ATA262000DTMA (2600)', 'ATADU2015 MTMA 1800&2100+2600 2IN-4OUT', 'ATA262000DTMA (2600)', 'DTMA1800 (no BOM)', 'DTMA1800&2100 (no BOM)']

	rfus_8 = ['LRFUe(800)']
	
	rfus_9 = ['MRFUd (900)', 'MRFU (900) old', 'MRFUv2 (900)'] 
	
	rfus_18 = ['MRFUv2 (1800)', 'MRFUd(1800)']
	
	rfus_21 = ['WRFUd(2100 2T2R) old', 'MARP 2100', 'WRFUd (2100 2T2R)'] 
	
	rru_18_4t4r = ['RRU3971 4T4R(1800)']
	
	# count TMAs
	
	tma_lb_count = 0
	tma_hb_count = 0
	rfus_8_count = 0
	rfus_9_count = 0
	rfus_18_count = 0
	rfus_21_count = 0
	rru_18_4t4r_count = 0
	lb_combiner_count = 0
	hb_combiner_count = 0
	r4t4_combiner_count = 0
	asi_antennas_count = 0
	
	# counting the TMAs and the ASI antennas
	for x in range(11): # counts the total number of tmas and distributes them in hb/lb lists. Also sums the ASIs.
		if sheet_ss.cell(column = 3, row = 53+x).value in tma_lb:
			tma_lb_count +=1
		elif sheet_ss.cell(column = 3, row = 53+x).value in tma_hb:
			tma_hb_count +=1
		elif x < 8 and sheet_ss.cell(column = 3, row = 45+x).value == 'ASI4518R11v06':
			asi_antennas_count +=1
	
	# counting the RFUs
	if sheet_ss.cell(column = 3, row = 34).value in rfus_8: # sums the 800 RFUs from column c
		rfus_8_count += int(sheet_ss.cell(column = 4, row = 34).value)
	if sheet_ss.cell(column = 6, row = 34).value in rfus_8: # sums the 800 RFUs from column f
		rfus_8_count += int(sheet_ss.cell(column = 7, row = 34).value)
	
	if sheet_ss.cell(column = 3, row = 35).value in rfus_9: # sums the 900 RFUs from column c
		rfus_9_count += int(sheet_ss.cell(column = 4, row = 35).value)
	if sheet_ss.cell(column = 6, row = 35).value in rfus_9: # sums the 900 RFUs from column f
		rfus_9_count += int(sheet_ss.cell(column = 7, row = 35).value)
	
	if sheet_ss.cell(column = 3, row = 36).value in rfus_18: # sums the 1800 RFUs from column c
		rfus_18_count += int(sheet_ss.cell(column = 4, row = 36).value)
	if sheet_ss.cell(column = 6, row = 36).value in rfus_18: # sums the 1800 RFUs from column f
		rfus_18_count += int(sheet_ss.cell(column = 7, row = 36).value)
	
	if sheet_ss.cell(column = 3, row = 37).value in rfus_21: # sums the 2100 RFUs from column c
		rfus_21_count += int(sheet_ss.cell(column = 4, row = 37).value)
	if sheet_ss.cell(column = 6, row = 37).value in rfus_21: # sums the 2100 RFUs from column f
		rfus_21_count += int(sheet_ss.cell(column = 7, row = 37).value)
	
	if sheet_ss.cell(column = 3, row = 36).value in rru_18_4t4r: # sums the 1800 4T4R RRUs from column c
		rru_18_4t4r_count += int(sheet_ss.cell(column = 4, row = 36).value)
	if sheet_ss.cell(column = 6, row = 36).value in rru_18_4t4r: # sums the 1800 4T4R RRUs from column f
		rru_18_4t4r_count += int(sheet_ss.cell(column = 7, row = 36).value)
	
	#evaluating the number of combiners
	if rfus_9_count + rfus_8_count != tma_lb_count and tma_lb_count != 0: # counting the 800&900 combiners
		lb_combiner_count = ( rfus_9_count + rfus_8_count ) - tma_lb_count
		if lb_combiner_count > rfus_9_count or lb_combiner_count > rfus_8_count: # fixing the cases where there is a sector without TMA
			lb_combiner_count = tma_lb_count

	
	if rfus_18_count + rfus_21_count != tma_hb_count and tma_hb_count != 0: # counting the 1800&2100 combiners
		hb_combiner_count = ( rfus_18_count + rfus_21_count ) - tma_hb_count
		if hb_combiner_count > rfus_18_count or hb_combiner_count > rfus_21_count: # fixing the cases where there is a sector without TMA
			hb_combiner_count = tma_hb_count
	
	
	if rru_18_4t4r_count > asi_antennas_count and asi_antennas_count != 0:  # counting the 1800&2100 combiners for 4t4r cases, when the number of 4t4r RRUs exceeds the number of ASI antennas
		r4t4_combiner_count = asi_antennas_count
	elif rru_18_4t4r_count != 0: # counting the 1800&2100 combiners for the rest of 4t4r cases
		r4t4_combiner_count = rru_18_4t4r_count
	
	
	# Writing to the dummy	
	if lb_combiner_count > 0:
		pos = 0
		for _ in range(lb_combiner_count) : # write the 800/900 combiners
			sheet_ss.cell(column = 3, row = 66+pos, value = 'ACOMD2H18 (800&900)') # Write the BBU for subrack 1
			sheet_ss.cell(column = 4, row = 66+pos, value = '1')
			pos +=2
	
	if hb_combiner_count > 0:
		pos = 0
		for _ in range(hb_combiner_count): # write the 1800/2100 combiners
			sheet_ss.cell(column = 3, row = 67+pos, value = 'ACOMD2H08 (1800&2100)') # Write the BBU for subrack 1
			sheet_ss.cell(column = 4, row = 67+pos, value = '1')
			pos +=2
	if r4t4_combiner_count > 0:
		pos = 0
		for _ in range(r4t4_combiner_count): # write the 1800/2100 combiners for 4t4r
			sheet_ss.cell(column = 6, row = 66+pos, value = 'ACOMD2H08 (1800&2100)') # Write the BBU for subrack 1
			sheet_ss.cell(column = 7, row = 66+pos, value = '2')
			pos +=2
	

	sswb.save(filename = dummy_file)
	


def add_dcstops(dummy_file):	
	
	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on	
	
	dcstops = 0
	for x in range(7): # counts the total number of tmas and distributes them in hb/lb lists. Also sums the ASIs.
		if sheet_ss.cell(column = 3, row = 66+x).value:
			dcstops += 2
	
	if dcstops:
			sheet_ss.cell(column = 3, row = 75, value = 'DC-Stop') # Write the BBU for subrack 1
			sheet_ss.cell(column = 4, row = 75, value = '{}'.format(dcstops))
	
	sswb.save(filename = dummy_file)
	
	
	
	
	
#def add_descriptors_existing(file, site_id, dummy_file):
#	ex_l8 = ''
#	ex_g9 = ''
#	ex_u9 = ''
#	ex_l9 = ''
#	ex_l18 = ''
#	ex_l18_4t4r = ''
#	ex_g18 = ''
#	ex_u21 = ''
#	ex_l21 = ''
#	ex_l26 = ''
#	ex_l26_4t4r = ''
#	
#	
#	# Parses technologies from the RNP export and builds the strings
#	with open(file, 'r', encoding ='utf-8') as csv_file:	# read from the RNP export
#		rnp_data = csv.reader(csv_file, delimiter=';')
#		for row in rnp_data:
#			if site_id in row[0] and row[5] == 'Active': # if the site id matches and the act status is 'Active'
#				if row[1][-1] == 'L':
#				
#				elif row[1][-1] == 'G'
#			
#				elif row[1][-1] == 'N'
#				
#				elif row[1][-1] == 'D'
#				
#				elif row[1][-1] == 'T'
#				
#				elif row[1][-1] == 'U'
#				
#				elif row[1][-1] == 'U' ( L26 )
#			
#			
#			
#			
#				if row[2]+'1' not in feeders.keys(): # add to dictionary with key A1, B1 or C1 key for first layer feeders
#					feeders[row[2]+'1'] = [row[20], row[19], row[4]]
#				elif row[2]+'2' not in feeders.keys(): # add to dictionary with key A2, B2 or C2 key for second layer feeders
#	
	
	
	
	
	
	
	
	


	
	
if __name__ == '__main__':
	print('Working. Please wait...')
	try:
		with open(output_files_list, 'r') as list:
			line = list.readline()
			while line:
				line = line.strip('\n')
				path = './output/'+line
				path_dummy = './output/dummy_'+line
				job_id = line[:10]
				shutil.copyfile(ss_template_file, path)
				shutil.copyfile(ss_template_file_dummy, path_dummy)
				rnd_shot(path_dummy, job_id)
				write_boards(parse_inv(inv_file, job_id[:5]),path_dummy)
				
				add_ant(rnp_summary_file, job_id[:5], path_dummy)
				add_ret(inv_antennas, job_id[:5], path_dummy)
				add_tma(inv_antennas, job_id[:5], path_dummy)
				add_cab(inv_cabinets, job_id[:5], path_dummy)
				add_feeder(rnp_summary_file, job_id[:5], path_dummy)
				add_owner_and_nis(nis_exp, job_id[:5], path_dummy)
				add_bbu(inv_subracks, job_id[:5], path_dummy)
				add_combiners(path_dummy)
				add_dcstops(path_dummy)
				
				line = list.readline()
	except:
		traceback.print_exc()
		input('Press "enter" to close the window')
