import shutil
import pandas as pd
from os import listdir
from os import rename
from datetime import datetime
from openpyxl import load_workbook
from keys_relation import direct_conversion_keys
from keys_relation import values_to_convert
from keys_relation import standard_value_dict
from keys_relation import values_to_compute
from keys_relation import cell_template_names
from keys_relation import rows_titles_all


rnp_template = 'RnpData_BTS3900 V100R015C00SPC100_13_44_24_empty_templ.xlsx'


# Parses the cell data provided by the Customer and returns a dataframe.
def parse(open_file_name):
    xls = pd.ExcelFile('./input/{}'.format(open_file_name))
    dtf = xls.parse(xls.sheet_names[2])

    return dtf.T.dropna()


# Adds the template first, as it's easier with non-converted data
def add_template(dat_fr):
    temps = {}
    for x in dat_fr.T.index:
        lst_int_par = ['_' + dat_fr[x]['*Cell transmission and reception mode'],
                       str(dat_fr[x]['*DlBandwidth']) + 'M',
                       dat_fr[x]['*Frequency Band']]

        for template_name in cell_template_names:
            if all(int_str in template_name for int_str in lst_int_par):
                temps[x] = template_name

    temps_df = pd.DataFrame.from_dict(temps, orient='index')
    temps_df.columns = ['CellTemplateName']
    dat_fr = pd.concat([dat_fr, temps_df.T])

    return dat_fr


# Converts the dicts keys to match the desired output format ones.
def convert(dat_fr):
    dat_fr.rename(index=direct_conversion_keys, inplace=True)  # Converts the line names
    dat_fr.drop([None, ], inplace=True)  # Removes the lines with no names after the conversion

    return dat_fr


# Updates the df to contain the necessary fields
def upd_log(dat_fr):

    # Adding a uplink bandwidth field by copying the Downlink bandwidth
    df_line = dat_fr.loc['Downlink bandwidth']
    dat_fr.rename(index={'Downlink bandwidth': 'Uplink bandwidth'}, copy=True, inplace=True)
    dat_fr = dat_fr.append(df_line)

    # Replacing the BW and Freq values
    dat_fr.T.replace(to_replace=values_to_convert, inplace=True)

    # Adding the Standard value fields ( There's a way to do this with a single line, I'm sure... have to learn pandas )
    standard_values_df = pd.DataFrame.from_dict(standard_value_dict, orient='index')
    standard_values_df = pd.concat([standard_values_df for _ in dat_fr.T.index], axis=1, ignore_index=False)
    standard_values_df = standard_values_df.T.reset_index()
    standard_values_df = standard_values_df.T
    dat_fr = dat_fr.append(standard_values_df, ignore_index=False)

    # Computes the remaining 2 based on the added info
    df_comp_values = pd.DataFrame()
    for x in dat_fr.T.index:
        df_comp_values = pd.concat([df_comp_values, pd.DataFrame.from_dict(values_to_compute.get(
            dat_fr[x]['*Cell transmission and reception mode']), orient='index')], ignore_index=False, axis=1)

    df_comp_values = df_comp_values.T.reset_index()
    df_comp_values = df_comp_values.T
    dat_fr = dat_fr.append(df_comp_values, ignore_index=False)

    # Drops the useless indexes added while concatting the frames
    dat_fr = dat_fr.drop(['index', ])

    return dat_fr


# Replaces the Danish characters if such exist
def char_rep(dat_fr):

    cor_names = pd.Series(dat_fr.T['*eNodeB Name']).str.replace('Ø', 'OE', regex=False)
    cor_names = pd.Series(cor_names).str.replace('Å', 'AA', regex=False)
    cor_names = pd.Series(cor_names).str.replace('Æ', 'AE', regex=False)
    dat_fr.replace(dat_fr.T['*eNodeB Name'], cor_names, inplace=True)

    return dat_fr


# Reindexing the columns so they match the output order
def reindex_col(dat_fr):

    dat_fr = dat_fr.reindex(index=rows_titles_all)

    return dat_fr


# Prints converted dicts to the output format.
def outprint(dat_fr):
    timestamp = datetime.now().strftime('%Y_%m_%d_%Hh_%Mm_%Ss')
    node_id = dat_fr[0]['*eNodeB ID']
    dest_file = './output/{0}_RnpData_{1}.xlsx'.format(node_id, timestamp)
    shutil.copyfile(rnp_template, dest_file)
    book = load_workbook(dest_file)
    writer = pd.ExcelWriter(dest_file, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    dat_fr.T.to_excel(writer,  index=False, sheet_name='Cell', startrow=1)
    writer.save()


# Reordering the sheets, so the CME check can pass - Currently out of use
def reorder(xls_file):
    wb = load_workbook(filename='./output/{}'.format(xls_file))
    myorder = [0, 1, 10, 2, 3, 4, 5, 6, 7, 8, 9]
    wb._sheets = [wb._sheets[i] for i in myorder]
    # wb.active = 3
    # wb["Prb"].views.sheetView[0].tabSelected = False
    wb.save('./output/{}'.format(xls_file))


# Adjusts the cell sheet so it can pass the CME verification
def final_adj(xls_file):
    wb = load_workbook(filename='./output/{}'.format(xls_file))
    sheet_cell = wb['Cell']
    sheet_cell.cell(column=1, row=1, value='eNodeB')
    sheet_cell.cell(column=3, row=1, value='Cell')
    sheet_cell.merge_cells('A1:B1')
    sheet_cell.merge_cells(start_row=1, start_column=3, end_row=1, end_column=66)

    out_file = './output/{}'.format(xls_file)
    wb.save(filename=out_file)

    rename(out_file, out_file[:-1])


if __name__ == '__main__':
    for file in listdir('./input/'):
        df = parse(file)
        df = add_template(df)
        df = convert(df)
        df = upd_log(df)
        df = char_rep(df)
        df = reindex_col(df)
        outprint(df)

        # print(df)  # Service print - to see the DataFrame

    for file in listdir('./output/'):
        # reorder(file) # Seems that CME doesn't care about the sheet order
        final_adj(file)
