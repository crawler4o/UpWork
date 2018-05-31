import shutil, csv
from openpyxl import load_workbook



rnd_file_name = 'Released plan B1, B2 2018.03.28.xlsx' 
output_files_list = 'list.txt'
ss_template_file = 'New Site Solution Template V3.8 Template.xlsx'
ss_template_file_dummy = 'Dummy_template.xlsx'
inv_file = 'Inventory_Board_20180328_175118.csv'
rnp_summary_file = 'MV_RNP_SUMMARY_20180326.csv'
inv_cabinets = 'Inventory_Cabinet_20180328_175133.csv'
inv_antennas = 'backup_antenna.csv'


def rnd_shot(ss, job_on): # This will paste the RND 
	wb = load_workbook(filename = rnd_file_name) # opening the rnd
	sswb = load_workbook(filename = ss) # opening the ss file
	sheet = wb['Running cells'] #defining the rnd sheet to read
	sheet_ss = sswb['Front Page'] # identifying the ss sheet to write on
	d_row = 3 # ss row to write the rnd data on
	for x in range(2, 3000):
		job = 'M'+ str(x) # row of the job in the rnd
		r_job = sheet[job].value # the job id from the rnd
		if job_on == r_job: # if the job name in the solution file name equals one on the rnd
			# print("MATCH!") # Troubleshooting entry - Prints match when a line corresponding to the current ss file is found.
			for col in range(1, 17):
				cp_s = sheet.cell(column = col, row = x ).value # source cell for copying
				if not cp_s:
					cp_s = ''
				sheet_ss.cell(column = col+7, row = d_row, value = "{0}".format(cp_s))		# writing the value to the destination cell
			d_row +=1 # ss going to write on the next row
	
	sswb.save(filename = ss)

		
def parse_inv(inv_file, site_id):
	bbu_boards = ['UPEU', 'UEIU', 'WBBP', 'LBBP', 'UBBP', 'UMPT', 'WMPT', 'LMPT', 'GTMU', 'UTRP'] # list of wanted boards to be kept in the BBU boards dictionary
	radio_boards = ['MRFU', 'WRFU', 'LRFU', 'MRRU', 'LRRU'] # list or boards to be kept in the RF boards dictionary
	
	with open(inv_file, 'r', encoding ='utf-8') as csv_file:
		brd_inv = csv.reader(csv_file)
		node = {}
		for row in brd_inv: # add the boards that correspond to the site
			if site_id in row[2]:
				node[row[23]] = [row[10], row[25], row[5], row[6]]
			
	to_del = [] # list of entries to be removed
	node_rf = {} # dictionary for the rf boards
	node_bbu = {}
	
	for x in node: # identify the boards
		if node[x][2] in radio_boards: # add the Rf boards to the node_rf dictionary
			node_rf[x] = node[x]
		
		if node[x][2] in bbu_boards:
			node_bbu[x] = node[x]
	
	node_bbu = convert_dic(node_bbu) # convert the dictionary for BBU
	node_rf = convert_dic(node_rf) # convert the dictionary for RF
	
	return [node_bbu, node_rf]
	

def write_boards(nodes, file): # writes the third field of the nodes dictionary ( now it's BOM, value 4 of the dict should be the RRU row ) 
	sswb = load_workbook(filename = file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on
	for x in nodes[0]: # checking and writing the BBU boards
		if nodes[0][x][0] == '0': # BBU in subrack 0
			if int(nodes[0][x][1]) < 8: # to write the boards on the first 7 slots
				if not sheet_ss.cell(column = 3, row = 22 +int(nodes[0][x][1])).value: # make sure that the field is empty
					sheet_ss.cell(column = 3, row = 22 +int(nodes[0][x][1]), value = "{0}".format(nodes[0][x][3]))
					sheet_ss.cell(column = 4, row = 22 +int(nodes[0][x][1]), value = 1)
				else:
					sheet_ss.cell(column = 3, row = 22 +int(nodes[0][x][1]), value = 'error')		
			else: #to write the UPEU and UEIU
				if not sheet_ss.cell(column = 3, row = 12 +int(nodes[0][x][1])).value: # make sure that the field is empty
					sheet_ss.cell(column = 3, row = 12 +int(nodes[0][x][1]), value = "{0}".format(nodes[0][x][3]))
					sheet_ss.cell(column = 4, row = 12 +int(nodes[0][x][1]), value = 1)
				else:
					sheet_ss.cell(column = 3, row = 12 +int(nodes[0][x][1]), value = 'error')
		else: # BBU in subrack 1
			if int(nodes[0][x][1]) < 8: # to write the boards on the first 7 slots
				if not sheet_ss.cell(column = 6, row = 22 +int(nodes[0][x][1])).value: # make sure that the field is empty
					sheet_ss.cell(column = 6, row = 22 +int(nodes[0][x][1]), value = "{0}".format(nodes[0][x][3]))
					sheet_ss.cell(column = 7, row = 22 +int(nodes[0][x][1]), value = 1)
				else:
					sheet_ss.cell(column = 6, row = 22 +int(nodes[0][x][1]), value = 'error')	
			else: #to write the UPEU and UEIU
				if not sheet_ss.cell(column = 6, row = 12 +int(nodes[0][x][1])).value: # make sure that the field is empty
					sheet_ss.cell(column = 6, row = 12 +int(nodes[0][x][1]), value = "{0}".format(nodes[0][x][3]))
					sheet_ss.cell(column = 7, row = 12 +int(nodes[0][x][1]), value = 1)
				else:
					sheet_ss.cell(column = 6, row = 12 +int(nodes[0][x][1]), value = 'error')

	for x in nodes[1]: # checking and writing the RF boards

		if not sheet_ss.cell(column = 3, row = int(nodes[1][x][4])).value: # if the field is empty
			sheet_ss.cell(column = 3, row = int(nodes[1][x][4]), value = "{0}".format(nodes[1][x][3]))
			sheet_ss.cell(column = 4, row = int(nodes[1][x][4]), value = 1)
		elif sheet_ss.cell(column = 3, row = int(nodes[1][x][4])).value == nodes[1][x][3]: #if the same is already added
			sheet_ss.cell(column = 4, row = int(nodes[1][x][4]), value = int(sheet_ss.cell(column = 4, row = int(nodes[1][x][4])).value) + 1)
		elif sheet_ss.cell(column = 3, row = int(nodes[1][x][4])).value != nodes[1][x][3] and not sheet_ss.cell(column = 6, row = int(nodes[1][x][4])).value: # if different in column 1 and nothing in column 2
			sheet_ss.cell(column = 6, row = int(nodes[1][x][4]), value = "{0}".format(nodes[1][x][3]))
			sheet_ss.cell(column = 7, row = int(nodes[1][x][4]), value = 1)
		elif sheet_ss.cell(column = 3, row = int(nodes[1][x][4])).value != nodes[1][x][3] and sheet_ss.cell(column = 6, row = int(nodes[1][x][4])).value == nodes[1][x][3]: # if different in column 1 and same in column 2
			sheet_ss.cell(column = 7, row = int(nodes[1][x][4]), value = int(sheet_ss.cell(column = 7, row = int(nodes[1][x][4])).value) + 1)
		else: # if different in both column 1 and 2 - return an error
			sheet_ss.cell(column = 3, row = int(nodes[1][x][4]), value = 'error')
				
	sswb.save(filename = file)

	
def convert_dic(dic): # convert the BOM codes into names as per the below dictionary. Adds row for the RF modules
	
	translate = { 
			'QWL3WBBPF3':['WBBPf3', '0'],
			'WD22UMPTb1':['UMPTb1', '0'],
			'WD2MUPEUC':['UPEUc', '0'],
			'WD5MJFUGG8E':['MRFUd  (900)', '35'],
			'QWL1WBBPD2':['WBBPd2', '0'],
			'WD22UMPTa2':['UMPTa2', '0'],
			'WD2M1UEIU':['UEIU', '0'],
			'WD5MJFUBG8E':['MRFUd  (900)', '35'],
			'QWL3WBBPF3':['WBBPf3', '0'],
			'WD22WMPT':['WMPT', '0'],
			'WD5MIFUBC10':['WRFUd (2100 2T2R)', '37'],
			'WD2MUPEUD2':['UPEUd', '0'],
			'WD5MJRUA880':['RRU3928 (900) old', '35'],
			'QWL1WBBPD1':['WBBPd1', '0'],
			'WD5MMRFU78':['MRFU (900) old', '35'],
			'WD22UBBPd1':['UBBPd1', '0'],
			'WD5MZAAZGAF':['3965d (800&900)', '34'],
			'WD22UBBPd4':['UBBPd4', '0'],
			'WD22UBBPd3':['UBBPd3', '0'],
			'WD5MZAAZGAFX':['RRU3965 (800&900)', '34'],
			'WD5MJRUE88E':['MRRU3938  (900)', '35'],
			'QWL1WBBPD3':['WBBPd3', '0'],
			'WD5MIRUD810':['RRU3838 2T2R(2100)', '37'],
			'QWM2UTRP4':['UTRP4', '0'],
			'WD22UMPTb2':['UMPTb2', '0'],
			'WD5MIRUA810':['RRU3828 2T2R(2100) old', '37'],
			'WD5MWFUB81':['MARP 2100', '37'],
			'WD5MJRUYCY0':['RRU 3961(800&900) old', '34'],
			'WD22UBBPd6':['UBBPd6', '0'],
			'WD2MWRFU81':['WRFUd(2100 2T2R) old', '37'],
			'QWL3WBBPF1':['WBBPf1', '0'],
			'QWL1WBBPF4':['WBBPf4', '0'],
			'WD5MMRFU78B':['MRFUv2 (900)', '35'],
			'WD3M1RRU4':['RRU3801E (2100)', '37'],
			'WD5MARU261':['RRU3804 (2100)', '37'],
			'WD5MIRUDC10':['RRU3839 2T2R(2100) new', '37'],
			'WD2MUPEUA':['UPEUa', '0'],
			'WD5MLFUHCK0':['LRFUe(800)', '34'],
			'WD5MJFUBG30':['MRFUd(1800)', '36'],
			'WD23LBBPD1':['LBBPd1', '0'],
			'WD5MIFUBCK0':['LRFUe(800)', '34'],
			'WD5MJFUGG30':['MRFUd(1800)', '36'],
			'WD22LMPT1':['LMPT', '0'],
			'WD22LBBPC':['LBBPc', '0'],
			'WD5MJFUHG30':['MRFUd(1800)', '36'],
			'WD5MIRUB8KA':['RRU3220 (800) old', '34'],
			'WD5MJRUA830':['RRU3928 (1800) old', '36'],
			'WD5MLRUH870':['RRU3268 2T2R(2600)', '38'],
			'WD5MJRUE830':['RRU3938(1800 2T2R)', '36'],
			'WD5MLRUA8K0':['LRRU3268(800)', '34'],
			'WD5MJRUIG30':['RRU3971 4T4R(1800)', '36'],
			'WD5MLRUYG70':['RRU3281 4T4R(2600)', '38'],
			'WD5MLFUHC70':['LRFU 2T2R(2600)', '38'],
			'WD5MLFU287C':['MARP(2600)', '38'],
			'WD5MLRUC870':['RRU3240(2600 2T4R) old', '38'],
			'WD5MLRUE870':['RRU3260(2600 2T4R) old', '38'],
			'WD22LBBPD1':['LBBPd1', '0'],
			'WD5MLRUA8K0L':['RRU3268 2T2R(2600)', '38'],
			'WD22LBBPD3':['LBBPd3', '0'],
			'WD22LBBPD2':['LBBPd2', '0'],
			'WD5MIRU187C':['RRU3201(2600) old', '38'],
			'WD5MMRFU73B':['MRFUv2 (1800)', '36'],
			'WD22GTMUb':['GTMUb', '0'],
		}
	
	for x in dic:
		dic[x].append(translate[dic[x][3]][1]) # adds the row 
		dic[x][3] = translate[dic[x][3]][0] # replaces the BOM with name based on 'board type'
		
	return dic

	
	
def add_ant(file, site_id, dummy_file): # adds the antennas to the dummy solution
	
	antennas = {} # dictionary to store the antennas.

	# Read the antennas from the RNP export and add the info to {}
	with open(file, 'r', encoding ='utf-8') as csv_file:	# read from the RNP export
		rnp_data = csv.reader(csv_file, delimiter=';')
		for row in rnp_data:
			if site_id in row[0]:
				if row[2] not in antennas.keys() or antennas[row[2]] == row[16]: # add to dictionary with clean A B or C key for first layer antennas
					antennas[row[2]] = row[16]
				elif antennas[row[2]] != row[16] and row[2]+'2' not in antennas.keys(): # add to dictionary with A2, B2, C2 for second layer antennas
					antennas[row[2]+'2'] = row[16]
				else: # adds 'error' value if more than 2 layers of antennas exist
					antennas[row[2]] = 'error'
	
		
	# Write the {} to the dummy solution
			
	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on	
		
	for x in antennas:
		if x == 'A':
			sheet_ss.cell(column = 3, row = 45, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 45, value = 1)
		elif x == 'B':
			sheet_ss.cell(column = 3, row = 47, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 47, value = 1)
		elif x ==  'C':
			sheet_ss.cell(column = 3, row = 49, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 49, value = 1)
		elif x == 'D':
			sheet_ss.cell(column = 3, row = 51, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 51, value = 1)
		elif x == 'A2':
			sheet_ss.cell(column = 3, row = 46, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 46, value = 1)
		elif x == 'B2':
			sheet_ss.cell(column = 3, row = 48, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 48, value = 1)
		elif x == 'C2':
			sheet_ss.cell(column = 3, row = 50, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 50, value = 1)
		elif x == 'D2':
			sheet_ss.cell(column = 3, row = 52, value = "{0}".format(antennas[x]))
			sheet_ss.cell(column = 4, row = 52, value = 1)

	sswb.save(filename = dummy_file)

	


##def add_cab(file, site_id, dummy_file)
##	cabinets = {} # The dictionary
##	with open(file, 'r', encoding ='utf-8') as csv_file: # read the cabinet export
##		cab_data = csv.reader(csv_file)
##		for row in cab_data:
##			if site_id in row[2]: # if the site ID is in the NE name
##				cabinets[row[15]] = row[14] # adds the value of 'Rack Type' with 'NS(BarCode) as a key
##	
##	for x in cabinets:
##		if 

	

def add_ret(file, site_id, dummy_file): # will use the 'Vendor Name' and 'Vendor Unit Family Type' fields to count + serial number as a key.
	rets = [] # The list
	with open(file, 'r', encoding ='utf-8') as csv_file: # read the antenna export
		ret_data = csv.reader(csv_file)
		for row in ret_data:
			if site_id in row[2] and row[19] == 'SINGLE_RET': # if the site ID is in the NE name
 				rets.append(row[18])

	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on	
	
	if rets.count('KA') > 0: 
		sheet_ss.cell(column = 4, row = 76, value = rets.count('KA'))
		sheet_ss.cell(column = 3, row = 76, value = "{0}".format('KA RET'))	
		
	sswb.save(filename = dummy_file)			
						


						
def add_tma(file, site_id, dummy_file):
	tmas = {} # key is the serial number
	with open(file, 'r', encoding = 'utf-8') as csv_file: # read the antenna export
		tma_data = csv.reader(csv_file)
		for row in tma_data:
			if site_id in row[2] and row[19] == 'TMA': # if the site ID is in the NE name
				if len(row[11]) > 2: # if the TMA has a BOM code 
					tmas[row[15]] = row[11] # the value is the BOM code
				else:
					tmas[row[15]] = row[14] # the value is the antenna model
	
		
	sswb = load_workbook(filename = dummy_file) # opening the dummy file
	sheet_ss = sswb['Site solution'] # identifying the ss sheet to write on		
	
	
	
	tma_translation_dic = { # Dictionary to convert the BOMs to names
					'27100072':'ATADU2001 (800&900)',
					'27100046':'STMA 2100',
					'27100074':'ATADU2002 (1800&2100)',
					'27100073':'ATADU2005 (800&900) 2IN-4OUT',
					'27100075':'ATADU2003 (1800+2100) 2IN-4OUT',
					'27100083':'ATADU2002 (1800&2100)',
					'27100045':'STMA 2100',
					'99044AHL':'STMA 2100',
					'27100060':'STMA 1800',
					'27100052':'ATA262000DTMA (2600)',
					'27100112':'ATADU2015 MTMA 1800&2100+2600 2IN-4OUT',
					'27100037':'STMA 1800',
					'DTMA2100':'STMA 2100',
					'78210517':'KA 78210517 ( 800 & 900 ,2in4out)',
					'DTMA800':'DTMA800 (no BOM)',
					'DTMA2600':'ATA262000DTMA (2600)',
					'DTMA800&900':'DTMA800&900 (no BOM)',
					'DTMA1800':'DTMA1800 (no BOM)',
					'DTMA1800&2100':'DTMA1800&2100 (no BOM)',
					}
	
	for x in tmas:
		tmas[x] = tma_translation_dic[tmas[x]] # convert what is on the TMA dict ( bom or ant model ) to the target values
	
	# filling the SS
	
	for x in tmas: # filling the SS
		pos1 = 53
		pos2 = 56
		pos3 = 59
		for z in range(2):
			if sheet_ss.cell(column = 3, row = pos1).value != tmas[x] and sheet_ss.cell(column = 3, row = pos1).value:
				pos1 +=1
				pos2 +=1
				pos3 +=1
			
		if sheet_ss.cell(column = 3, row = pos1).value == tmas[x]: # check if the same is already filled in somewhere
			if sheet_ss.cell(column = 3, row = pos2).value == tmas[x]:
				if sheet_ss.cell(column = 3, row = pos3).value == tmas[x]:
					sheet_ss.cell(column = 3, row = pos3, value = "{0}".format('error'))
				else:
					sheet_ss.cell(column = 3, row = pos3, value = "{0}".format(tmas[x]))
					sheet_ss.cell(column = 4, row = pos3, value = 1)
			else:
				sheet_ss.cell(column = 3, row = pos2, value = "{0}".format(tmas[x]))
				sheet_ss.cell(column = 4, row = pos2, value = 1)
		elif not sheet_ss.cell(column = 3, row = pos1).value: # if it's empty
			sheet_ss.cell(column = 3, row = pos1, value = "{0}".format(tmas[x]))
			sheet_ss.cell(column = 4, row = pos1, value = 1)			
	
	
	sswb.save(filename = dummy_file)						

						
if __name__ == '__main__':
	print('Working. Please wait...')
	with open(output_files_list, 'r') as list:
		line = list.readline()
		while line:
			line = line.strip('\n')
			path = './output/'+line
			path_dummy = './output/dummy_'+line
			job_id = line[:10]
			shutil.copyfile(ss_template_file, path)
			shutil.copyfile(ss_template_file_dummy, path_dummy)
			rnd_shot(path_dummy, job_id)
			write_boards(parse_inv(inv_file, job_id[:5]),path_dummy)
			add_ant(rnp_summary_file, job_id[:5], path_dummy)
			add_ret(inv_antennas, job_id[:5], path_dummy)
			add_tma(inv_antennas, job_id[:5], path_dummy)
			line = list.readline()
